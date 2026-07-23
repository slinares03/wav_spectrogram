"""Signal detection over a spectrogram, and CSV logs for Excel / Sheets.

Three outputs, because "when a frequency is hit" means different things
depending on what you are doing with the numbers:

  detections  one row per peak per time slice — the raw firehose. Good for
              plotting or pivoting; a 3-minute file can be tens of thousands
              of rows.
  events      contiguous detections merged into one row per signal: start,
              end, duration, frequency span. This is usually the one you
              actually want to read.
  watch       one row per time slice for a frequency *you* name, whether or
              not anything was detected there. Use it to answer "was my
              1.2 kHz tone present, and how strong?"

Detection rule
--------------
A bin counts as hit when it sits `snr_db` above the noise floor *of its own
time slice* (the median across frequency in that row).

Taking the floor per row rather than per bin is deliberate: a per-bin floor
computed across time is raised by any steady tone until the tone is level with
its own threshold and stops being detected. A steady carrier is exactly what
this is for, so the floor has to come from the other frequencies present at the
same moment. The median (not the mean) keeps a few loud bins from dragging the
floor up and masking the quieter signals beside them.
"""

from __future__ import annotations

import csv
from pathlib import Path

import numpy as np


# Excel on Windows guesses the encoding of a plain CSV and mangles non-ASCII
# (the Hz/dB column headers survive, but a filename with an accent will not)
# unless the file starts with a BOM. utf-8-sig writes one; Google Sheets and
# LibreOffice both skip it silently, so this is safe everywhere.
_ENCODING = "utf-8-sig"


def _prominence(row: np.ndarray, peak: int) -> float:
    """How far `peak` stands above the higher of the two saddles beside it.

    Look outward until the trace rises above the peak (or the band ends), and
    take the lowest point reached on each side. The higher of those two minima
    is the "key col"; prominence is the peak's height above it. This is what
    separates a genuine second tone from a shoulder on a loud one's skirt —
    both are local maxima, but only the tone stands clear of the valley
    between them.

    Done with numpy slicing rather than a Python walk. An isolated strong tone
    has no higher neighbour, so the walk ran the full band: profiling a
    3-minute file showed 9.0M interpreter-level min() calls and 8.3 s of a
    9.0 s analysis inside this one function. The result here is identical —
    the search is bounded the same way, just executed in C.
    """
    h = row[peak]
    left = row[:peak]
    if left.size:
        higher = np.flatnonzero(left > h)
        start = higher[-1] + 1 if higher.size else 0
        lo_left = left[start:].min() if start < left.size else h
    else:
        lo_left = h
    right = row[peak + 1:]
    if right.size:
        higher = np.flatnonzero(right > h)
        stop = higher[0] if higher.size else right.size
        lo_right = right[:stop].min() if stop > 0 else h
    else:
        lo_right = h
    return float(h - max(lo_left, lo_right))


def _bandwidth(row: np.ndarray, peak: int, drop_db: float = 3.0):
    """Index span where the peak stays within `drop_db` of its maximum."""
    limit = row[peak] - drop_db
    a = peak
    while a > 0 and row[a - 1] >= limit:
        a -= 1
    b = peak
    while b < row.size - 1 and row[b + 1] >= limit:
        b += 1
    return a, b


def detect(psd, freqs, times, snr_db: float = 12.0,
           prominence_db: float = 6.0, min_separation_hz: float = 0.0):
    """Find spectral peaks in a (time, freq) dB matrix.

    Returns one dict per peak per time slice.

    Peaks, not runs
    ---------------
    An earlier version treated each contiguous over-threshold span as one
    detection. That collapses under a loud tone: a 110 dB-SNR carrier holds ~50
    bins above a 12 dB threshold purely through spectral leakage, so its span
    merges with every other signal it reaches and the whole slice reports as a
    single 2 kHz-wide blob. Picking local maxima and filtering them by
    prominence keeps tones separate no matter how wide their skirts are.
    """
    out = []
    bin_hz = float(freqs[1] - freqs[0]) if freqs.size > 1 else 1.0
    min_sep = max(int(round(min_separation_hz / bin_hz)), 1)
    floors = np.median(psd, axis=1)                 # noise floor per time slice
    for i, row in enumerate(psd):
        floor = floors[i]
        # Local maxima, tall enough to matter. The >= on one side keeps flat
        # two-bin tops from being missed entirely.
        hi = (row[1:-1] > row[:-2]) & (row[1:-1] >= row[2:]) & (row[1:-1] >= floor + snr_db)
        cand = np.flatnonzero(hi) + 1
        if cand.size == 0:
            continue
        # Strongest first, so when two peaks are closer than the minimum
        # separation the weaker (usually a skirt shoulder) is the one dropped.
        taken = []
        check_sep = min_sep > 1          # at the default there is nothing to check
        for p in cand[np.argsort(row[cand])[::-1]]:
            prom = _prominence(row, int(p))     # keep it: recomputing per output
            if prom < prominence_db:            # row doubled this function's cost
                continue
            if check_sep and any(abs(int(p) - q) < min_sep for q, _ in taken):
                continue
            taken.append((int(p), prom))
        for p, prom in sorted(taken):
            a, b = _bandwidth(row, p)
            seg = row[a:b + 1]
            # Power-weighted centroid over the -3 dB span: for a tone straddling
            # two bins this lands between them rather than snapping to whichever
            # bin happens to be larger.
            w = 10.0 ** (seg / 10.0)
            # Rounded at source: these land in a spreadsheet, where
            # 0.9600000000000001 is noise, and it lets the bulk sheets skip
            # per-cell number formatting (see write_xlsx).
            out.append({
                "time_s": round(float(times[i]), 6),
                "freq_hz": round(float(freqs[p]), 3),
                "centroid_hz": round(float(np.sum(freqs[a:b + 1] * w) / np.sum(w)), 3),
                "amplitude_db": round(float(row[p]), 2),
                "snr_db": round(float(row[p] - floor), 2),
                "prominence_db": round(prom, 2),
                "bandwidth_hz": round(float((b - a + 1) * bin_hz), 3),
                "freq_low_hz": round(float(freqs[a]), 3),
                "freq_high_hz": round(float(freqs[b]), 3),
                "_row": i,
            })
    return out


def group_events(dets, max_gap_s: float = 0.15, freq_tol_hz: float = 0.0,
                 min_duration_s: float = 0.0):
    """Merge detections that are close in time and frequency into events.

    Two detections join the same event when their frequency spans overlap (with
    `freq_tol_hz` of slack) and they are within `max_gap_s` of each other. The
    gap tolerance matters: a real tone drops below threshold for the odd frame
    when it fades or a louder signal lifts the floor, and without it a single
    steady carrier fragments into hundreds of one-frame events.
    """
    active, done = [], []
    for d in sorted(dets, key=lambda x: x["time_s"]):
        lo, hi = d["freq_low_hz"] - freq_tol_hz, d["freq_high_hz"] + freq_tol_hz
        hit = None
        for ev in active:
            if lo <= ev["freq_high_hz"] and hi >= ev["freq_low_hz"] \
                    and d["time_s"] - ev["end_s"] <= max_gap_s:
                hit = ev
                break
        if hit is None:
            active.append({
                "start_s": d["time_s"], "end_s": d["time_s"],
                "freq_low_hz": d["freq_low_hz"], "freq_high_hz": d["freq_high_hz"],
                "peak_db": d["amplitude_db"], "peak_freq_hz": d["freq_hz"],
                "_sum": d["centroid_hz"], "_n": 1, "_snr": d["snr_db"],
            })
        else:
            hit["end_s"] = d["time_s"]
            hit["freq_low_hz"] = min(hit["freq_low_hz"], d["freq_low_hz"])
            hit["freq_high_hz"] = max(hit["freq_high_hz"], d["freq_high_hz"])
            hit["_sum"] += d["centroid_hz"]
            hit["_n"] += 1
            hit["_snr"] = max(hit["_snr"], d["snr_db"])
            if d["amplitude_db"] > hit["peak_db"]:
                hit["peak_db"] = d["amplitude_db"]
                hit["peak_freq_hz"] = d["freq_hz"]
        # Retire events nothing has extended recently, so the active list stays
        # short on long files instead of being rescanned in full every row.
        still = [e for e in active if d["time_s"] - e["end_s"] <= max_gap_s]
        done.extend(e for e in active if e not in still)
        active = still
    done.extend(active)

    out = []
    # Number *after* the duration filter, so ids are always 1..n with no holes.
    kept = [e for e in sorted(done, key=lambda x: x["start_s"])
            if e["end_s"] - e["start_s"] >= min_duration_s]
    for i, e in enumerate(kept, start=1):
        dur = e["end_s"] - e["start_s"]
        out.append({
            "event_id": i,
            "start_s": round(e["start_s"], 6),
            "end_s": round(e["end_s"], 6),
            "duration_s": round(dur, 6),
            "center_hz": round(e["_sum"] / e["_n"], 3),
            "peak_freq_hz": round(e["peak_freq_hz"], 3),
            "freq_low_hz": round(e["freq_low_hz"], 3),
            "freq_high_hz": round(e["freq_high_hz"], 3),
            "bandwidth_hz": round(e["freq_high_hz"] - e["freq_low_hz"], 3),
            "peak_db": round(e["peak_db"], 2),
            "max_snr_db": round(e["_snr"], 2),
            "n_detections": e["_n"],
        })
    return out


def watch(psd, freqs, times, targets, snr_db: float = 12.0,
          tol_hz: float | None = None):
    """Track named frequencies over time: one row per target per time slice.

    `tol_hz` averages power over a band around each target instead of reading a
    single bin — useful when the tone drifts or falls between bins. Defaults to
    one bin either side.
    """
    bin_hz = float(freqs[1] - freqs[0]) if freqs.size > 1 else 1.0
    tol = bin_hz if tol_hz is None else float(tol_hz)
    floors = np.median(psd, axis=1)
    rows = []
    for f in targets:
        sel = np.abs(freqs - float(f)) <= tol
        if not sel.any():                       # target outside the band
            sel = np.zeros(freqs.size, bool)
            sel[int(np.argmin(np.abs(freqs - float(f))))] = True
        actual = float(freqs[sel].mean())
        band = psd[:, sel].max(axis=1)          # strongest bin in the band
        for i, t in enumerate(times):
            snr = float(band[i] - floors[i])
            rows.append({
                "time_s": round(float(t), 6),
                "target_hz": float(f),
                "measured_hz": round(actual, 3),
                "amplitude_db": round(float(band[i]), 2),
                "noise_floor_db": round(float(floors[i]), 2),
                "snr_db": round(snr, 2),
                "present": int(snr >= snr_db),
            })
    return rows


EVENT_COLUMNS = ["event_id", "start_s", "end_s", "duration_s", "center_hz",
                 "peak_freq_hz", "freq_low_hz", "freq_high_hz", "bandwidth_hz",
                 "peak_db", "max_snr_db", "n_detections"]
DETECTION_COLUMNS = ["time_s", "freq_hz", "centroid_hz", "amplitude_db",
                     "snr_db", "prominence_db", "bandwidth_hz", "freq_low_hz",
                     "freq_high_hz"]
WATCH_COLUMNS = ["time_s", "target_hz", "measured_hz", "amplitude_db",
                 "noise_floor_db", "snr_db", "present"]


def build_sheets(psd, freqs, times, snr_db=12.0, prominence_db=6.0,
                 min_separation_hz=0.0, max_gap_s=0.15, min_duration_s=0.0,
                 watch_hz=None, watch_tol=None, include_detections=True):
    """Run the full analysis and return [(title, rows, columns), ...].

    One place builds the tables, so the CLI and the player's export button
    cannot drift into producing different columns from the same file.
    """
    dets = detect(psd, freqs, times, snr_db=snr_db,
                  prominence_db=prominence_db,
                  min_separation_hz=min_separation_hz)
    events = group_events(dets, max_gap_s=max_gap_s,
                          min_duration_s=min_duration_s)
    sheets = [("Events", events, EVENT_COLUMNS)]
    if include_detections:
        sheets.append(("Detections", dets, DETECTION_COLUMNS))
    if watch_hz:
        rows = watch(psd, freqs, times, watch_hz, snr_db=snr_db, tol_hz=watch_tol)
        sheets.append(("Watch", rows, WATCH_COLUMNS))
    return sheets


# Number formats by column suffix, so seconds don't render as 0.9600000000001
# and frequencies don't collapse to scientific notation on an RF capture.
_FORMATS = (("_s", "0.000"), ("_hz", "0.0"), ("_db", "0.0"))
# Above this many rows a sheet is bulk data, not something read by eye.
_FORMAT_ROW_LIMIT = 5000

# Workbook look, matched to WAV_Spectrogram_Test_Plan.xlsx: a Carlito face, a
# dark-navy banner/header with white text, and a pale-blue accent for values.
_XLSX_FONT = "Carlito"
_XLSX_NAVY = "FF17365D"       # banner + header fill / accent text
_XLSX_PALE = "FFEAF3F8"       # accent cell fill
_XLSX_WHITE = "FFFFFFFF"


def xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def write_xlsx(path: Path, sheets, info=None) -> dict:
    """Write one workbook with a sheet per table.

    `sheets` is a sequence of (title, rows, columns). Returns {title: n_rows}.

    Preferred over CSV for this because the three tables are one artefact, not
    three loose files, and because a spreadsheet stores real numeric types —
    a CSV is text that Excel re-guesses on open, which is where "1.0E+09" and
    date-mangled values come from.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    navy_fill = PatternFill("solid", fgColor=_XLSX_NAVY)
    pale_fill = PatternFill("solid", fgColor=_XLSX_PALE)
    banner_font = Font(name=_XLSX_FONT, bold=True, size=16, color=_XLSX_WHITE)
    header_font = Font(name=_XLSX_FONT, bold=True, color=_XLSX_WHITE)
    accent_font = Font(name=_XLSX_FONT, bold=True, color=_XLSX_NAVY)
    body_font = Font(name=_XLSX_FONT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    banner_align = Alignment(vertical="center")

    # Vertical rhythm copied from the test plan: a tall banner (row 1), a blank
    # spacer row (row 2), then a tall wrapped header (row 3), then data. Keeping
    # HDR_ROW in one place is what keeps the freeze/filter/format offsets honest.
    HDR_ROW = 3
    BANNER_H, HEADER_H = 30.0, 34.0

    def banner(ws, text, ncols):
        """Row 1: a merged navy title banner spanning the table's width."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = ws.cell(row=1, column=1, value=text)
        cell.font = banner_font
        cell.fill = navy_fill
        cell.alignment = banner_align
        ws.row_dimensions[1].height = BANNER_H
        for c in range(1, ncols + 1):           # fill spans the whole merge
            ws.cell(row=1, column=c).fill = navy_fill
        ws.sheet_view.showGridLines = False

    def header(ws, cols):
        """Bold white-on-navy header cells at HDR_ROW, with a spacer above."""
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=HDR_ROW, column=c, value=name)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = header_align
        ws.row_dimensions[HDR_ROW].height = HEADER_H

    wb = Workbook()
    wb.remove(wb.active)                        # drop the default empty sheet
    counts = {}

    if info:
        ws = wb.create_sheet("Run info")
        banner(ws, "WAV Spectrogram Analysis", 2)
        header(ws, ["setting", "value"])
        for r, (k, v) in enumerate(info.items(), start=HDR_ROW + 1):
            kc = ws.cell(row=r, column=1, value=k)
            vc = ws.cell(row=r, column=2, value=v)
            kc.font = body_font
            vc.font = accent_font
            vc.fill = pale_fill
            vc.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 52
        ws.freeze_panes = f"A{HDR_ROW + 1}"

    for title, rows, columns in sheets:
        ws = wb.create_sheet(title)
        cols = list(columns)
        banner(ws, title, len(cols))
        header(ws, cols)
        for i, r in enumerate(rows, start=HDR_ROW + 1):
            for c, name in enumerate(cols, start=1):
                ws.cell(row=i, column=c, value=r.get(name))

        # Freeze the banner+header and switch on filters: with tens of thousands
        # of detection rows, scrolling without these is unusable.
        ws.freeze_panes = f"A{HDR_ROW + 1}"
        last = get_column_letter(len(cols))
        if rows:
            ws.auto_filter.ref = f"A{HDR_ROW}:{last}{HDR_ROW + len(rows)}"
        # Number formats cost one attribute write per cell — ~0.4 s of a 1.4 s
        # save on a 3-minute file. Apply them to the small human-read sheets and
        # skip them on the bulk ones, whose values are already rounded at source
        # and so display correctly under the default format anyway.
        style_cells = len(rows) <= _FORMAT_ROW_LIMIT
        for i, name in enumerate(cols, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = max(len(name) + 3, 11)
            if not style_cells:                 # bulk sheet: skip per-cell work
                continue
            fmt = next((f for suf, f in _FORMATS if name.endswith(suf)), None)
            for cell in ws[letter][HDR_ROW:]:   # body rows, past the header
                cell.font = body_font
                if fmt:
                    cell.number_format = fmt
        counts[title] = len(rows)

    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    wb.save(path)
    return counts


def write_csv(rows, path: Path, columns=None) -> int:
    """Write dict rows to a CSV. Returns the number of data rows written."""
    path = Path(path)
    if not rows:
        # Still write the header, so a downstream script or a spreadsheet
        # formula sees an empty table rather than a missing file.
        cols = list(columns or [])
    else:
        cols = list(columns or [k for k in rows[0] if not k.startswith("_")])
    with path.open("w", newline="", encoding=_ENCODING) as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return len(rows)
